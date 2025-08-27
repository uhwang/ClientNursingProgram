from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QDate
import cnpdb, cnputil, cnpval

def save_client_to_excel(clients, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Info"

    # --- Names ---
    ws.append(["Type", "First", "Middle", "Last"])
    ws.append(["Korean", cnputil.safe_val(clients.get(cnpdb.col_first_name_kor)), "",
                          cnputil.safe_val(clients.get(cnpdb.col_last_name_kor))])
    ws.append(["English", cnputil.safe_val(clients.get(cnpdb.col_first_name_eng)),
                          cnputil.safe_val(clients.get(cnpdb.col_middle_name_eng)),
                          cnputil.safe_val(clients.get(cnpdb.col_last_name_eng))])
    ws.append([])

    # --- DOB/AGE/SEX ---
    qdate_obj = QDate.fromString(clients.get(cnpdb.col_dob,''), cnpval.date_format_r)
    cur_age = cnputil.calculate_age(qdate_obj)
    
    ws.append(["DOB", "Age", "Sex"])
    ws.append([cnputil.safe_val(clients.get(cnpdb.col_dob)),
               f"{cur_age}",
               cnputil.safe_val(clients.get(cnpdb.col_sex))])
    ws.append([])

    # --- Assessments ---
    ws.append(["Assessment Type", "Date"])
    ws.append(["Initial", cnputil.safe_val(clients.get(cnpdb.col_initial_assessment))])
    ws.append(["14th", cnputil.safe_val(clients.get(cnpdb.col_assessment_14th))])
    ws.append(["90th", cnputil.safe_val(clients.get(cnpdb.col_assessment_90th))])
    ws.append(["Change", cnputil.safe_val(clients.get(cnpdb.col_change_assessment))])
    ws.append(["Done", cnputil.safe_val(clients.get(cnpdb.col_change_assessment_done))])
    ws.append([])

    # --- Room & Comments ---
    ws.append(["Room Number", cnputil.safe_val(clients.get(cnpdb.col_room_number))])
    ws.append(["Comments", cnputil.safe_val(clients.get(cnpdb.col_comments))])

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(filepath)

def export_clients_to_excel(clients, visible_columns, filepath, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title 
    ws.append(visible_columns)
    
    # ---- Normalize rows (dict) and string-ify, handling None ----
    for r_ in clients:
        # sqlite3.Row → dict; dict stays as-is
        record = []
        for c_ in visible_columns:
            key = cnpdb.available_view_column_key(c_)
            val = r_.get(key, "")
            if val is None:
                val = ""
            record.append(val)
        ws.append(record)
        
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(filepath)
    