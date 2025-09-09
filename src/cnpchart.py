'''
    cnpchart.py
    9/8/2025
    AI : gemini

'''
from collections import defaultdict
#import matplotlib.pyplot as plt
#import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import PatternFill
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties # Correct import

from PyQt5.QtWidgets import (
     QPushButton   , QHBoxLayout   , QVBoxLayout, 
     QDialog      , QFormLayout   , QComboBox,
     QGridLayout, QCheckBox, QLabel
)

from PyQt5.QtCore import Qt, QSize, QDate
from PyQt5.QtGui import QPixmap, QIcon

import msg, color
import cnpdb, cnpval, cnpconf, cnpvtbl, msg, cnpval, cnputil
import icon_system, icon_view_setting, icon_chart_bar, \
       icon_chart_pie
      
_bar_chart_fname = "Age-bar-chart.xlsx"      
_pie_chart_fname = "Age-pie-chart.xlsx"
_age_interval = 5
      
class QChartDlg(QDialog):
    def __init__(self, db_man, gmsg):
        super(QChartDlg, self).__init__()
        self.db_man = db_man
        self.client_list = None
        self.gmsg = gmsg
        m_layout = QGridLayout()
        
        self.bar_chart = QCheckBox()
        self.bar_chart.setIcon(QIcon(QPixmap(icon_chart_bar.table)))
        self.bar_chart.setIconSize(QSize(32,32))
        
        self.pie_chart = QCheckBox()
        self.pie_chart.setIcon(QIcon(QPixmap(icon_chart_pie.table)))
        self.pie_chart.setIconSize(QSize(32,32))
        
        self.ok = QPushButton('OK')
        self.cancel = QPushButton('CANCEL')
        
        self.ok.clicked.connect(self.create_chart)
        self.cancel.clicked.connect(self.reject)
        
        m_layout.addWidget(self.bar_chart, 0,0)
        m_layout.addWidget(QLabel("Individual Age"), 0,1)
        m_layout.addWidget(self.pie_chart, 1,0)
        m_layout.addWidget(QLabel("Age Distribution"), 1,1)
        m_layout.addWidget(self.ok, 2,0)
        m_layout.addWidget(self.cancel, 2,1)
        self.setLayout(m_layout)
        self.setWindowTitle("Chart")
        self.setWindowIcon(QIcon(QPixmap(icon_system.table)))

    def create_chart(self):
        if not any([self.bar_chart.isChecked(), self.pie_chart.isChecked()]):
            msg.message_box("Chose at least one option")
            return
        
        key = [cnpdb.col_id, cnpdb.col_dob]
        
        try:
            clients = self.db_man.db.load_all_clients(key)
        except Exception as e:
            e_msg = f"cnpchart.py (create_chart): {e}"
            self.gmsg.appendPlainText(e_msg)
            msg.message_box(e_msg)
            return
        
        if clients == []:
            e_msg = "cnpchart.py (create_chart): No clients!!"
            self.gmsg.appendPlainText(e_msg)
            msg.message_box(e_msg)
            return
        
        client_data = []
        for c_ in clients:
            i_ = c_[cnpdb.col_id][-3:]
            d_ = c_[cnpdb.col_dob][0:10] # get only MM/dd/yyyy
            new_date = QDate.fromString(d_, cnpval.date_format_w)
            client_data.append([i_, cnputil.calculate_age(new_date)])
            
        if self.bar_chart.isChecked():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Individual Age"
                
                for row_data in client_data:
                    ws.append(row_data)
    
                # --- 2. Create a Bar Chart for Client Ages ---
                # Create a reference to the data for the bar chart
                data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(client_data))
                
                # Create a reference to the categories (client names)
                labels_ref = Reference(ws, min_col=1, min_row=1, max_row=len(client_data))
            
                # Create the bar chart object
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = "Client Ages"
                bar_chart.y_axis.title = "Age"
                bar_chart.x_axis.title = "Client"
                bar_chart.y_axis.visible = True
                bar_chart.y_axis.majorTickMark = "in"
                bar_chart.y_axis.tickLblPos = 'nextTo'
                # Customize the appearance of bars
            
                # Add the data and categories to the chart
                bar_chart.add_data(data_ref, titles_from_data=True)
                bar_chart.set_categories(labels_ref)
                bar_chart.series[0].graphicalProperties.solidFill = "4472C4"  # Blue color
                bar_chart.series[0].graphicalProperties.line.solidFill = "2E5A87"  # Darker blue border
            
                # Place the chart on the worksheet starting at cell E2
                ws.add_chart(bar_chart, "E2")            
                wb.save(_bar_chart_fname)
                wb = None
            except Exception as e:
                e_msg = f"Error: Pie Chart\n{e}"
                self.gmsg.appendPlainText(e_msg)
                msg.message_box(e_msg)
                return
            
        if self.pie_chart.isChecked():
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Individual Age"
                    
                ages = [a_[1] for a_ in client_data]
                # Use a defaultdict to simplify counting
                age_groups = defaultdict(int)
            
                # Find the minimum age to determine the starting group
                min_age = min(ages)
                
                # Calculate the starting point for the first group
                start_group = (min_age // _age_interval) * _age_interval
            
                # Find the maximum age to determine the final group
                max_age = max(ages)
                
                # Iterate through all ages and count them into their respective groups
                for age in ages:
                    # Determine the start of the 5-year _age_interval for the current age
                    group_start = (age // _age_interval) * _age_interval
                    group_end = group_start + _age_interval - 1
                    
                    # Create the group label (e.g., "20-24")
                    group_label = f"{group_start}-{group_end}"
                    
                    # Increment the count for this group
                    age_groups[group_label] += 1
                
                age_groups = {group: count for group, count in age_groups.items() if count > 0}

                # --- Write age group data to the worksheet ---
                # The chart needs to reference data in cells, so we write it first.
                ws.cell(row=1, column=1, value="Age Group")
                ws.cell(row=1, column=2, value="Count")
        
                # Start writing data from the second row
                for i, (group, count) in enumerate(age_groups.items(), start=2):
                    ws.cell(row=i, column=1, value=group)
                    ws.cell(row=i, column=2, value=count)

                # --- Create the Pie Chart ---
                pie_chart = PieChart()
                pie_chart.title = "Client Age Group Distribution"
                
                # Define the data and labels for the chart
                # We need to reference the cells where we wrote the data
                data_ref = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
                labels_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                
                # 1. Generate colors using matplotlib's 'viridis' colormap
                num_pieces = len(age_groups)
                #colormap = plt.colormaps['tab20c']
                #colormap = plt.colormaps['RdYlGn']
                #colormap = plt.colormaps['hsv']
                #colormap = plt.colormaps['Spectral']
                colormap = color.create_color_table(0,190,0.5,1,ws.max_row)
                #colors = colormap(np.linspace(0, 1, num_pieces))
                #hex_colors = [f'#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}' for r, g, b, a in colors]
                hex_colors = [f'#{c_.r:02x}{c_.g:02x}{c_.b:02x}' for c_ in colormap]

                # Create a DataLabelList object for customization
                data_labels = DataLabelList()
                # Set show_percent to True to display percentages
                data_labels.show_percent = True
                # Set the position to 'bestFit' to prevent labels from overlapping
                data_labels.position = 'bestFit'
                pie_chart.data_labels = data_labels
                
                # Create the series and add it to the chart's series list
                series = Series(values=data_ref)
                series.dLbls = DataLabelList()
                series.dLbls.showCatName = True
                series.dLbls.showVal = True
                series.dLbls.showPercent = True
                series.dLbls.showSerName = False
                series.dLbls.showLegendKey = False
                series.dLbls.position = 'bestFit' 
                pie_chart.series.append(series)
                
                # 5. Apply the custom colors to each slice
                series_ = pie_chart.series[0]
                
                for i, color_hex in enumerate(hex_colors):
                    color_choice = ColorChoice(prstClr=None, srgbClr=color_hex.lstrip('#'))
                    # Create a GraphicalProperties object
                    gp = GraphicalProperties(solidFill=color_choice)
                    
                    # Assign the GraphicalProperties object to DataPoint's spPr
                    dp = DataPoint(idx=i, spPr=gp)
                    series_.dPt.append(dp)
                    
                # The corrected line: directly append the series to the chart's series list
                pie_chart.set_categories(labels_ref)

                # Add the chart to the worksheet
                # The chart will be positioned starting at cell D2
                ws.add_chart(pie_chart, "D2")
                wb.save(_pie_chart_fname)
                wb = None
            except Exception as e:
                e_msg = f"Error: Pie Chart\n{e}"
                self.gmsg.appendPlainText(e_msg)
                msg.message_box(e_msg)
                return
        self.accept()
        
def create_client_chart(db_man, gmsg):
    
    b_ = QChartDlg(db_man, gmsg)
    ret = b_.exec_()
